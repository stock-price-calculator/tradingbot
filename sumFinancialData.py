from openpyxl import Workbook, load_workbook
import pandas as pd

# Company that you find
file = 'samsung.xlsx'

wb = load_workbook(file)

# 재무상태표 시트 지정
ws_bs = wb['Data_bs']

# 년도 가져오기
y_list = []
c = 10
while ws_bs.cell(2, c).value != None:
    y = ws_bs.cell(1, c).value
    if y[:1] == '2':
        y_list.append(y)
    c += 1

# 가져올 항목 List로 만들기
bs_items = ['ifrs-full_Assets', 'ifrs-full_CurrentAssets', 'ifrs-full_CashAndCashEquivalents',
            'dart_ShortTermTradeReceivable',
            'entity00126380_udf_BS_201710182279121_CurrentAssets', 'ifrs-full_Inventories', 'ifrs-full_Liabilities',
            'ifrs-full_CurrentLiabilities', 'ifrs-full_ShorttermBorrowings', 'ifrs-full_Equity',
            'ifrs-full_IssuedCapital',
            'ifrs-full_RetainedEarnings']

# 재무상태표 Data 가져오기
df_list = []
for bs_item in bs_items:
    temp_list = []
    r = 4
    while ws_bs.cell(r, 2).value != None:
        if ws_bs.cell(r, 2).value == bs_item:
            item = ws_bs.cell(r, 3).value
            c = 10
            while ws_bs.cell(1, c).value != None:
                temp_list.append(ws_bs.cell(r, c).value)
                c += 1
        r += 1
    df = pd.DataFrame({item: temp_list}, index=y_list)
    df_list.append(df)

total_df = pd.concat(df_list, axis=1)
total_df = total_df.transpose()

print(total_df)
total_df.to_excel('재무상태표.xlsx')